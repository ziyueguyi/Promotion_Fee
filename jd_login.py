# -*- coding:utf-8 -*-
# @文件名称  :jd_login.py
# @项目名称  :Promotion_Fee
# @软件名称  :PyCharm
# @创建时间  :2021-04-06 10:54
# @用户名称  :紫月孤忆

import os
import shutil

import cv2
import json
import time
import random
import asyncio
import pandas as pd
import win32api
import win32con
import threading
import numpy as np
import multiprocessing

from urllib import request

from colorama import Fore, Style
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyppeteer import launch
from base_fun import funtion, mso
from base_fun import send_message
from datetime import datetime, date
from funtions.jd_jzt_cost_bk import JztCost
from apscheduler.schedulers.blocking import BlockingScheduler

# wx_key = 'b9c7e2d4-f972-454d-9db9-ae2b27cdb38d'
wx_key = None
send = send_message.Send(wx_key=wx_key)


async def main(un, pw, sn):
    width = win32api.GetSystemMetrics(win32con.SM_CXSCREEN)  # 获得屏幕分辨率X轴
    height = win32api.GetSystemMetrics(win32con.SM_CYSCREEN)  # 获得屏幕分辨率Y轴
    browser = await launch({
        'headless': False,
        'userDataDir': './tool/datas/{0}/userdata'.format(sn),
        'dumpio': True,
        'args': ['--disable-infobars',
                 '--no-sandbox',
                 '--window-size={0},{1}'.format(width, height),
                 ],
    })
    statue = None
    page = await browser.newPage()
    shop_name = Fore.LIGHTGREEN_EX + '{0}'.format(sn) + Style.RESET_ALL
    try:
        await page.setUserAgent(
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4542.2 Safari/537.36')
        await page.setViewport({'width': width, 'height': height})
        page.setDefaultNavigationTimeout(1000 * 20)  # 设置默认等待时间
        await page.goto(
            'https://passport.jd.com/common/loginPage?from=jzt&amp;ReturnUrl=https%3A%2F%2Fjzt.jd.com%2Fhome%2F%23%2Findex',
            {'timeout': 10000 * 20, 'waitUntil': 'networkidle0'})
        time.sleep(random.randint(3, 6))
        await page.evaluateOnNewDocument(
            '''() =>{ Object.defineProperties(navigator, { webdriver: { get: () => false } }) }''')
        try:
            await page.type('#loginname', un,
                            {'delay': random.randint(60, 121)})
            await page.type('#nloginpwd', pw,
                            {'delay': random.randint(100, 151)})
            await page.click('#paipaiLoginSubmit')
            time.sleep(random.randint(3, 6))
            await page.goto('https://jzt.jd.com/home/#/index',
                            {'timeout': 10000 * 20, 'waitUntil': 'networkidle0'})
        except BaseException as e:
            print(e, 6)
        await page.evaluateOnNewDocument(
            '''() =>{ Object.defineProperties(navigator, { webdriver: { get: () => false } }) }''')
        time.sleep(random.randint(3, 6))
        cookies = await page.cookies()
        path = r'./tool/datas/{0}/cookies/{1}_{2}.json'.format(sn, 'jd', sn)

        statue = Fore.LIGHTBLUE_EX + "cookies获取成功" + Style.RESET_ALL
        try:
            funtion.jud_path(path, flag=False)
            with open(path, 'w') as f:
                json.dump(cookies, f)
        except BaseException as e:
            statue = Fore.LIGHTRED_EX_EX + "cookies获取失败" + Style.RESET_ALL
            ''.format(e)
        await page.waitFor(5000)
    except BaseException as e:
        print(e, 5)
        statue = Fore.LIGHTRED_EX_EX + "浏览器模拟失败" + Style.RESET_ALL
    finally:
        print('{0}:{1}'.format(shop_name, statue))
        await page.close()
        await browser.close()


async def s_ver(page, frame, image1, image2, sn):
    """
    模拟验证
    :param sn:
    :param image1: 待验证的大图
    :param image2: 验证的小图
    :param page:浏览器窗口
    :param frame:
    :return:
    """
    # 获取两个需要验证的图片
    i1, i2 = image1, image2
    image_src = await frame.Jeval('.JDJRV-bigimg >img', 'el => el.src')
    request.urlretrieve(image_src, i1)
    template_src = await frame.Jeval('.JDJRV-smallimg >img', 'el => el.src')
    request.urlretrieve(template_src, i2)
    await page.waitFor(3000)
    # 获取需要验证的控件
    el = await frame.J('div.JDJRV-slide-btn')
    box = await el.boundingBox()
    await frame.hover('div.JDJRV-slide-btn')
    distance = await get_distance(i1, i2)
    print('获取距离'.format(sn))
    await page.mouse.down()
    await page.mouse.move(box['x'] + distance + random.uniform(30, 33), box['y'], {'steps': 30})
    await page.waitFor(random.randint(300, 700))
    await page.mouse.move(box['x'] + distance + 29, box['y'], {'steps': 30})
    await page.mouse.up()
    await page.waitFor(3000)
    print('验证结束')


async def get_distance(i1, i2, types=True, dv=278 / 360):
    """
    滑块的缺口距离识别
    :param dv: 图片距离偏差值
    :param types: 读取的图片是否携带中文路径
    :param i1:需要验证的大图路径
    :param i2:验证的小图片路径
    :return:
    """
    if types:
        # 读取带中文路径图片的方法
        img = cv2.imdecode(np.fromfile(i1, dtype=np.uint8), -1)
        template = cv2.imdecode(np.fromfile(i2, dtype=np.uint8), -1)
    else:
        # 读取不带中文路径图片的方法
        img = cv2.imread(i1, 0)
        template = cv2.imread(i2, 0)
    res = cv2.matchTemplate(img, template, cv2.TM_CCORR_NORMED)
    value = cv2.minMaxLoc(res)[2][0]
    distance = value * dv
    return distance


def r_fun(username, password, shop_name, name, jc, mgr):
    # try:
    asyncio.get_event_loop().run_until_complete(main(username, password, shop_name))
    time.sleep(1)
    lock = threading.Lock()
    r_data = jc.start(name, shop_name)
    lock.acquire()
    try:
        # 放心地改吧:
        mgr.append(r_data)
    finally:
        lock.release()


def g_sql(fn, mgr):
    if mgr:
        sku_type = ['kc', 'ht', 'tp']
        shop_user = ['shop_name', 'user_name']
        xt_sku_type = ['x_kc', 'x_ht', 'x_tp', 'x_hz']
        ht_sku_type = ['y_kc', 'y_ht', 'y_tp', 'y_hz']
        bc_sku_type = ['c_kc', 'c_ht', 'c_tp', 'c_hz']
        xt_sku_type_n = ['xn_kc', 'xn_ht', 'xn_tp', 'xn_hz']
        data = data_jh('jd_ztc_cost', [shop_user, sku_type, xt_sku_type])[[*shop_user, *xt_sku_type]]
        n_data = data_jh('jd_ztc_cost_none', [shop_user, sku_type, xt_sku_type_n])[[*shop_user, *xt_sku_type_n]]
        data.rename(columns=dict(zip(sku_type, xt_sku_type)), inplace=True)
        if n_data.shape[0]:
            data = pd.merge(data, n_data, on=shop_user, how='left')
            data.fillna(0, inplace=True)
            for i in range(0, len(xt_sku_type)):
                data[xt_sku_type[i]] = data[xt_sku_type[i]] + data[xt_sku_type_n[i]]
            data = data[[*shop_user, *xt_sku_type]]
        pfd_pd = pd.DataFrame(list(mgr), columns=[*shop_user, *ht_sku_type])
        pfd_pd['y_hz'] = 0
        new_data = pd.merge(pfd_pd, data, how='left', on=shop_user)
        new_data = new_data.fillna(0)
        for i in xt_sku_type + ht_sku_type:
            new_data[i] = pd.to_numeric(new_data[i], errors='coerce')
        for i in range(0, len(bc_sku_type)):
            new_data[bc_sku_type[i]] = pd.to_numeric(new_data[ht_sku_type[i]] - new_data[xt_sku_type[i]],
                                                     errors='coerce')
        new_data = new_data.round(2)
        all_columns = shop_user + xt_sku_type + ht_sku_type + bc_sku_type
        new_data = new_data[all_columns]
        zw_columns = ['店铺', '负责人',
                      '系统快车金额', '系统海投金额', '系统触点金额', '系统汇总金额',
                      '应算快车金额', '应算海投金额', '应算触点金额', '应算汇总金额',
                      '补差快车金额', '补差海投金额', '补差触点金额', '补差汇总金额']
        new_data.rename(
            columns=dict(zip(all_columns, zw_columns)), inplace=True)
        new_data = new_data.style.applymap(
            lambda x: 'background-color: #6495ED' if x > 1 else 'background-color: #FFFFFF',
            subset=['补差快车金额', '补差海投金额', '补差触点金额', '补差汇总金额'])
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            new_data.to_excel(writer, sheet_name='京东推广费补差', index_label=None, index=None)
        reset_col(fn)
        time.sleep(1)
        send.send_file(fn)
    else:
        print('未成功推送')


def reset_col(filename):
    """
    计算最适合的列宽，并设置到excel里面
    :param filename:
    :return:
    """
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 0.8  # 也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)


def data_jh(db_name, columns):
    s_str = """SELECT shop_name,user_name,sku_type,SUM(sku_cost) FROM {0} WHERE  create_time>=(SELECT DATE_FORMAT(NOW(),'%Y-%m-%d 00:00:00') AS '今天开始') GROUP BY shop_name,user_name,sku_type""".format(
        db_name)
    data = mso.bing_mysql(s_str, db_type='pro', tip=False)
    sku_money = ['sku_type', 'sum_money']
    dt = pd.DataFrame(data, columns=[*columns[0], *sku_money])
    if dt.shape[0]:
        dt.set_index(columns[0], inplace=True)
        dt = dt.reset_index().pivot(columns[0], *sku_money)
        dt.reset_index(inplace=True)
        dt.rename(columns=dict(zip(columns[1], columns[2])), inplace=True)
        for i in columns[2]:
            if i not in dt.columns:
                dt[i] = 0
    else:
        dt = pd.DataFrame(columns=[*columns[0], *columns[2]])
    return dt


# def g_ski(nd):
#     pass


def g_sku_non(fn):
    """
    获取没有匹配到sku的推广费信息
    :param fn:
    :return:
    """
    s_sql = """select sku_id,sku_cost,sku_code,sku_type,shop_name,create_date,user_name from jd_ztc_cost_none"""
    data = mso.bing_mysql(s_sql, db_type='pro', tip=False)
    dayOfWeek = datetime.now().isoweekday()
    if data and dayOfWeek not in (6, 7):
        data = pd.DataFrame(data, columns=['商品编号', '花费', '商品sku', '推广费类型', '店铺名称', '日期', '责任人'])
        data.to_excel(fn, index_label=None, index=None)
        time.sleep(1)
        send.send_file(fn)
        send.send_msg('请各位负责人抓紧时间填写对应sku')


def run(num=5, flag=True, shop_name=None):
    """
    多进程运行程序，默认为五个进程
    :param shop_name:
    :param flag: 多进程运营，还是单进程运行
    :param num:
    :return:
    """
    route = './tool/'
    jc = JztCost()
    db = mso.DC(db_type='pro')
    s_sql = 'select user_name,pwd,shop_name,name from jd_jzt_user where use_off="1" and shop_type = 1'
    jd_user = db.bing_mysql(s_sql)
    shop_name = [i for i in jd_user if flag or i[2] in shop_name]
    print([sn[2] for sn in shop_name])
    if shop_name:
        mgr = multiprocessing.Manager().list()
        pool = multiprocessing.Pool(processes=num)
        [pool.apply_async(r_fun, (*i, jc, mgr)) for i in shop_name]  # 将账号放入线程池，等待数据获取，默认限制五个线程
        pool.close()
        pool.join()
        send_file(route, jc, mgr, jd_user)
    else:
        send.send_msg('没有查询到可用账号')
        print('没有查询到可用账号')


def send_file(route, jc, mgr, jd_user):
    try:
        dt = datetime.now().strftime('%Y-%m-%d')
        filename = route + '{0}/{1}/{2}/补差文档/'.format(*jc.get_route(dt))
        funtion.chect_dir(filename)
        g_sql(filename + '补差文档{0}.xlsx'.format(dt), mgr)
        g_sku_non(filename + '京东推广费{0}.xlsx'.format(dt))
        remove_ushop(jd_user, route)
        remove_over_data(route)
    except BaseException as e:
        print(e, 4)
        send.send_msg('京东推广费推送失败')


def remove_ushop(jd_user, route):
    """
    移除文件夹下无用的店铺信息，防止文件过多
    :param jd_user:
    :param route:
    :return:
    """
    shop_list = [ju[2] for ju in jd_user]
    route += 'datas/'
    shop_dir = os.listdir(route)
    for sd in shop_dir:
        if sd not in shop_list:
            try:
                shutil.rmtree(route + sd)
            except BaseException as e:
                print('移除无用的店铺信息失败', e)


def g_three_month(mm):
    year = date.today().year
    month = date.today().month - mm
    if month <= 0:
        year -= 1
        month += 12
    return year, month


def remove_over_data(rt):
    route_list = [os.path.join(rt, 'datas').replace('\\', '/')]
    year_list = [rt]
    for i in range(0, 3):
        year, month = g_three_month(i)
        month = os.path.join(rt, str(year), str(month)).replace('\\', '/')
        year = os.path.join(rt, str(year)).replace('\\', '/')
        year_list.append(year)
        route_list.extend([year, month])
    year_list = list(set(year_list))
    route_list = list(set(route_list))
    for yl in year_list:
        lds = os.listdir(yl)
        if 'tm_data' in lds:
            lds.remove('tm_data')
        for ld in lds:
            rts = os.path.join(yl, ld).replace('\\', '/')
            if rts not in route_list and os.path.exists(rts):
                shutil.rmtree(rts)


@funtion.add_time
def main_run(flag=True):
    one_start = ['头号卖家官方旗舰店_3', '头号卖家官方旗舰店_2']
    run(flag=flag, shop_name=one_start)


def test():
    fn = './tool/2021/7/30/京东.xlsx'
    c = [['头号卖家官方旗舰店', '王宁', 1, 2, 3, 0]]
    g_sql(fn, c)


if __name__ == '__main__':
    # test()
    # exit()
    print('程序启动')
    # 单次执行
    main_run(flag=False)
    # 定时运行
    bs = BlockingScheduler()
    bs.add_job(main_run, 'cron', hour=8, minute=0, misfire_grace_time=1000 * 90)
    bs.start()
